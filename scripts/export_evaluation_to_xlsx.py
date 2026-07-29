#!/usr/bin/env python3
"""
export_evaluation_to_xlsx.py — Flatten all 450 evaluation JSON files into a single XLSX.

Each row = one run (CVE × model × run_number).
Each column = one flattened key from the JSON (dot-separated path).

Usage:
    python export_evaluation_to_xlsx.py [--eval-root /path/to/evaluation] [--output output.xlsx]
"""
import argparse
import glob
import json
import os
from collections import OrderedDict


def flatten(obj, prefix="", out=None):
    """Recursively flatten a nested dict/list into dot-separated keys."""
    if out is None:
        out = OrderedDict()
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else k
            flatten(v, key, out)
    elif isinstance(obj, list):
        # For lists of dicts, expand each item with index
        for idx, item in enumerate(obj):
            key = f"{prefix}[{idx}]"
            flatten(item, key, out)
    else:
        out[prefix] = obj
    return out


def extract_run_data(json_path):
    """Extract a flat dict of all key-value pairs from one evaluation JSON."""
    with open(json_path) as f:
        data = json.load(f)
    return flatten(data)


def main():
    ap = argparse.ArgumentParser(description="Export evaluation JSONs to XLSX.")
    ap.add_argument("--eval-root", default="/home/tca/reprobench/eval/evaluation",
                    help="Root directory containing CVE-*/model/run/evaluation-*.json")
    ap.add_argument("--output", default="evaluation_results.xlsx",
                    help="Output XLSX file path")
    args = ap.parse_args()

    # Find all JSON files
    json_files = sorted(glob.glob(os.path.join(args.eval_root,
                                                "CVE-*", "*", "*",
                                                "evaluation-*.json")))
    print(f"Found {len(json_files)} JSON files")

    # Extract data from each file
    rows = []
    all_keys = set()

    for jf in json_files:
        # Parse path: .../CVE-XXXX/model/run/evaluation-CVE-XXXX.json
        parts = jf.split(os.sep)
        cve = None
        model = None
        run_num = None
        for p in parts:
            if p.startswith("CVE-"):
                cve = p
        # Model and run are the two directories before the JSON
        model = parts[-3]
        run_num = int(parts[-2])

        flat = extract_run_data(jf)
        flat["cve"] = cve
        flat["model"] = model
        flat["run"] = run_num
        flat["json_path"] = jf
        rows.append(flat)
        all_keys.update(flat.keys())

    # Define column order: metadata first, then plan_score, task_score, overall, failure
    # Put cve/model/run first, then sort the rest
    priority_cols = ["cve", "model", "run"]
    # Collect all keys and sort: priority first, then alphabetical
    remaining = sorted(all_keys - set(priority_cols))
    columns = priority_cols + remaining

    # Write to XLSX
    try:
        from openpyxl import Workbook
    except ImportError:
        import subprocess
        subprocess.check_call(["pip3", "install", "openpyxl"])
        from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            # Truncate very long strings (evidence, summaries)
            if isinstance(value, str) and len(value) > 32767:
                value = value[:32767]
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width (approximate)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, min(len(rows) + 2, 100)):  # Sample first 100 rows
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    # Freeze header row and first 3 columns
    ws.freeze_panes = "D2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(columns)).column_letter}{len(rows) + 1}"

    wb.save(args.output)
    print(f"Written: {args.output}")
    print(f"  Rows: {len(rows)}")
    print(f"  Columns: {len(columns)}")


if __name__ == "__main__":
    main()
